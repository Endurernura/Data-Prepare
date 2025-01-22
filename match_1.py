import openpyxl

#为name匹配ATC/DBID
def match_names():
    # 打开atc.xlsx文件
    atc_workbook = openpyxl.load_workbook(r'D:\Personal\Desktop\DDI_PROject\最终打包数据（不可用）\ComInfo\Target\Raw\Target_Lung.xlsx')
    atc_sheet = atc_workbook['CDCDB']

    # 打开Name.xlsx文件
    name_workbook = openpyxl.load_workbook(r'D:\Personal\Desktop\DDI_PROject\DB_ID与特征\DB_Target\DB_Target.xlsx')
    name_sheet = name_workbook.active

    # 创建一个字典，用于存储Name.xlsx中DBid与Name的对应关系. 第0列是DBID，第1列是Target。
    name_dict = {}
    for row in name_sheet.iter_rows(min_row=1, values_only=True):
        name_dict[row[1]] = row[0]

    # 遍历atc.xlsx的每一行数据(从第二行开始)
    for row in atc_sheet.iter_rows(min_row=2, values_only=False):
        dbid1 = row[5].value
        if dbid1 in name_dict:
            row[4].value = name_dict[dbid1]



    # 保存修改后的atc.xlsx文件
    atc_workbook.save(r'D:\Personal\Desktop\DDI_PROject\最终打包数据（不可用）\ComInfo\Target\Raw\lt3.xlsx')


if __name__ == '__main__':
    match_names()
