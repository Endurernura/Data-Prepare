import openpyxl

def convert_columns_to_rows():
    # 打开Excel文件
    workbook = openpyxl.load_workbook(r'D:\Personal\Desktop\ALL DATA\Similarity\ATC\ATC_Breast_1.xlsx')
    sheet = workbook['AIO']

    # 获取第一列、第二列、第三列的数据
    column1_data = [cell.value for cell in sheet['A'][1:]]
    column2_data = [cell.value for cell in sheet['B'][1:]]
    column3_data = [cell.value for cell in sheet['C'][1:]]

    # 清空第一行、第二行、第三行的数据（除了第一列的表头，假设第一列是表头）
    for row in sheet['A1:C3']:
        for cell in row:
            if cell.column!= 'A':
                cell.value = None

    # 将第一列的数据依次填入第一行的第二列到最后一列
    for i, value in enumerate(column1_data, start=2):
        sheet.cell(row=1, column=i).value = value

    # 将第二列的数据依次填入第二行的第二列到最后一列
    for i, value in enumerate(column2_data, start=2):
        sheet.cell(row=2, column=i).value = value

    # 将第三列的数据依次填入第三行的第二列到最后一列
    for i, value in enumerate(column3_data, start=2):
        sheet.cell(row=3, column=i).value = value

    # 保存修改后的Excel文件
    workbook.save(r'D:\Personal\Desktop\ALL DATA\Similarity\ATC\ATC_Breast_2.xlsx')


if __name__ == '__main__':
    convert_columns_to_rows()
