import openpyxl
from rdkit import Chem
from rdkit.Chem import AllChem
from rdkit.DataStructs import FingerprintSimilarity

def calculate_similarities(input_file_path, output_file_path):
    """
    计算Excel文件中指定区域的SMILES值之间的相似性，并将结果保存到新的Excel文件中
    """
    # 加载输入的Excel文件
    workbook = openpyxl.load_workbook(input_file_path)
    sheet = workbook.active

    # 获取行数和列数
    max_row = sheet.max_row
    max_column = sheet.max_column  # 将max_col修改为max_column

    # 创建一个新的工作簿用于保存结果
    result_workbook = openpyxl.Workbook()
    result_sheet = result_workbook.active

    # 写入表头，第一行第一列留空，后续依次写入列标题和行标题对应的SMILES
    result_sheet.cell(row=1, column=1, value="")
    for col_index in range(2, max_column + 1):  # 将max_col修改为max_column
        result_sheet.cell(row=1, column=col_index, value=sheet.cell(row=1, column=col_index).value)
    for row_index in range(2, max_row + 1):
        result_sheet.cell(row=row_index, column=1, value=sheet.cell(row=row_index, column=1).value)

    # 遍历行和列，计算相似性并填充结果到新的Excel表中
    for row_index in range(2, max_row + 1):
        for col_index in range(2, max_column + 1):  # 将max_col修改为max_column
            smiles_1 = sheet.cell(row=row_index, column=1).value
            smiles_2 = sheet.cell(row=1, column=col_index).value

            mol_1 = Chem.MolFromSmiles(smiles_1)
            mol_2 = Chem.MolFromSmiles(smiles_2)

            if mol_1 is None or mol_2 is None:
                similarity = "无效SMILES"
            else:
                fp_1 = AllChem.GetMorganFingerprint(mol_1, 2)
                fp_2 = AllChem.GetMorganFingerprint(mol_2, 2)
                similarity = FingerprintSimilarity(fp_1, fp_2)

            result_sheet.cell(row=row_index, column=col_index, value=similarity)

    # 保存结果工作簿
    result_workbook.save(output_file_path)


if __name__ == "__main__":
    input_file_path = r"D:\Personal\Desktop\DDI_PROject\OUTPUT\FINAL_Prepare\SMILES\rtcb.xlsx"  # 替换为实际的输入文件路径
    output_file_path = r"D:\Personal\Desktop\DDI_PROject\OUTPUT\FINAL_Prepare\SMILES\SMILES_Breast.xlsx"  # 替换为实际的输出文件路径
    calculate_similarities(input_file_path, output_file_path)

