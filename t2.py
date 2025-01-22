from openpyxl import load_workbook


def jaccard_similarity(set1, set2):
    """
    计算两个集合的Jaccard相似性
    """
    intersection = set1 & set2
    union = set1 | set2
    return len(intersection) / len(union) if union else 0


# 加载Excel文件
workbook = load_workbook(r'D:\Personal\Desktop\DDI_PROject\最终打包数据（不可用）\Similarity\Target\rtcb2.xlsx')
sheet = workbook['Target']

# 创建一个新的工作表用于存放结果
result_sheet = workbook.create_sheet(title="Results")

# 获取第一行数据（除第一列单元格），并处理成集合形式，去除重复BEID
first_row_sets = []
for cell in sheet[1][1:]:
    elements = cell.value.split(',') if cell.value else []
    element_set = set([element.strip() for element in elements])
    first_row_sets.append(element_set)

# 获取第一列数据（除第一行单元格），并处理成集合形式，去除重复BEID
first_column_sets = []
for row in sheet.iter_rows(min_row=2, max_col=1):
    cell = row[0]
    elements = cell.value.split(',') if cell.value else []
    element_set = set([element.strip() for element in elements])
    first_column_sets.append(element_set)

# 计算并将Jaccard相似性输出到对应的单元格中
for col_idx, col_set in enumerate(first_row_sets):
    for row_idx, row_set in enumerate(first_column_sets):
        similarity = jaccard_similarity(col_set, row_set)
        # 在新工作表中对应位置写入相似性结果，行列索引都从1开始（符合Excel单元格坐标习惯）
        result_sheet.cell(row=row_idx + 2, column=col_idx + 2, value=similarity)

# 保存修改后的Excel文件
workbook.save(r'D:\Personal\Desktop\DDI_PROject\tg1.xlsx')