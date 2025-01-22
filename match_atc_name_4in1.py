import openpyxl

#为name匹配ATC/DBID
def match_names():
    # 打开atc.xlsx文件
    atc_workbook = openpyxl.load_workbook(r'D:\Personal\Desktop\DDI_PROject\IEEE JBHI 2024Winter\All You Need Is\ComInfo\ATC\Raw\ATC_Breast_1.xlsx')
    atc_sheet = atc_workbook['CDCDB']

    # 打开Name.xlsx文件
    name_workbook = openpyxl.load_workbook(r'D:\Personal\Desktop\DDI_PROject\IEEE JBHI 2024Winter\DB_ID与特征\DB_ATC&Name.xlsx')
    name_sheet = name_workbook.active

    # 创建一个字典，用于存储Name.xlsx中DBid与Name的对应关系. 第0列是DBID，第1列是ATC。
    name_dict = {}
    for row in name_sheet.iter_rows(min_row=1, values_only=True):
        name_dict[row[0]] = row[1]
    # 遍历atc.xlsx的每一行数据(从第二行开始)
    for row in atc_sheet.iter_rows(min_row=2, values_only=False):
        dbid1 = row[13].value
        if dbid1 in name_dict:
            row[14].value = name_dict[dbid1]
        dbid2 = row[16].value
        if dbid2 in name_dict:
            row[17].value = name_dict[dbid2]
        dbid3 = row[19].value
        if dbid3 in name_dict:
            row[20].value = name_dict[dbid3]
        dbid4 = row[22].value
        if dbid4 in name_dict:
            row[23].value = name_dict[dbid4]


    # 保存修改后的atc.xlsx文件
    atc_workbook.save(r'D:\Personal\Desktop\DDI_PROject\IEEE JBHI 2024Winter\All You Need Is\ComInfo\ATC\Raw\ATC_Breast.xlsx')


if __name__ == '__main__':
    match_names()
