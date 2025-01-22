import openpyxl
from rdkit import Chem
from rdkit.Chem import RDKFingerprint
from rdkit.DataStructs import FingerprintSimilarity

def calculate_similarities(input_file_path, output_file_path):

    workbook = openpyxl.load_workbook(input_file_path)
    sheet = workbook['SMILES']

    max_row = sheet.max_row
    max_column = sheet.max_column

    result_workbook = openpyxl.Workbook()
    result_sheet = result_workbook.active

    result_sheet.cell(row=1, column=1, value="")
    for col_index in range(2, max_column + 1):
        result_sheet.cell(row=1, column=col_index, value=sheet.cell(row=1, column=col_index).value)
    for row_index in range(2, max_row + 1):
        result_sheet.cell(row=row_index, column=1, value=sheet.cell(row=row_index, column=1).value)

    for row_index in range(2, max_row + 1):
        for col_index in range(2, max_column + 1):
            smiles_1 = sheet.cell(row=row_index, column=1).value
            smiles_2 = sheet.cell(row=1, column=col_index).value

            mol_1 = Chem.MolFromSmiles(smiles_1)
            mol_2 = Chem.MolFromSmiles(smiles_2)

            fp_1 = RDKFingerprint(mol_1)
            fp_2 = RDKFingerprint(mol_2)
            similarity = FingerprintSimilarity(fp_1, fp_2)

            result_sheet.cell(row=row_index, column=col_index, value=similarity)


    result_workbook.save(output_file_path)

#rt1处理三象限的数据。

if __name__ == "__main__":
    input_file_path = r"D:\Personal\Desktop\DDI_PROject\OUTPUT\FINAL_Prepare\SMILES\rtcl1.xlsx"
    output_file_path = r"D:\Personal\Desktop\DDI_PROject\OUTPUT\FINAL_Prepare\SMILES\rtcl1_1.xlsx"    calculate_similarities(input_file_path, output_file_path)

