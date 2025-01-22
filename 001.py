import openpyxl
import multiprocessing
import math
import collections

def process_rows(start_row, end_row, input_file_path, output_file_path, shared_data):
    """
    处理指定行范围的数据，查找并删除第一列值重复的行（只保留重复值中最小行号对应的行），
    并将本进程处理范围内出现的第一列数据及对应的最小行号更新到共享数据结构中
    """
    workbook = openpyxl.load_workbook(input_file_path)
    sheet = workbook.active
    local_first_column_min_rows = {}
    rows_to_delete = []

    for row_index in range(start_row, end_row + 1):
        row = list(sheet.iter_rows(min_row=row_index, max_row=row_index, values_only=True))[0]
        if len(row) >= 1:
            first_column_value = row[0]
            if first_column_value in local_first_column_min_rows:
                if local_first_column_min_rows[first_column_value] > row_index:
                    rows_to_delete.append(local_first_column_min_rows[first_column_value])
                    local_first_column_min_rows[first_column_value] = row_index
                else:
                    rows_to_delete.append(row_index)
            else:
                local_first_column_min_rows[first_column_value] = row_index

    # 根据要删除的行号删除行
    for row_index in sorted(rows_to_delete, reverse=True):
        sheet.delete_rows(row_index)

    # 使用共享数据结构，将本进程处理范围内的数据合并到总的数据中
    for key, value in local_first_column_min_rows.items():
        shared_data[key] = min(shared_data[key], value) if key in shared_data else value

    workbook.save(output_file_path)


def process_sheet_data_parallel():
    input_file_path = r'C:\Users\17492\000.xlsx'
    output_file_path = r'C:\Users\17492\000_processed.xlsx'
    workbook = openpyxl.load_workbook(input_file_path)
    sheet = workbook.active
    total_rows = sheet.max_row
    num_processes = 8

    rows_per_process = math.ceil(total_rows / num_processes)
    manager = multiprocessing.Manager()
    shared_data = manager.dict()

    processes = []
    for i in range(num_processes):
        start_row = i * rows_per_process + 2
        end_row = min((i + 1) * rows_per_process + 1, total_rows)
        p = multiprocessing.Process(target=process_rows, args=(start_row, end_row, input_file_path, output_file_path, shared_data))
        processes.append(p)
        p.start()

    for p in processes:
        p.join()

    # 最后根据共享数据结构中的最小行号，再次遍历整个文件删除其他重复行
    workbook = openpyxl.load_workbook(output_file_path)
    sheet = workbook.active
    rows_to_delete_final = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) >= 1:
            first_column_value = row[0]
            if row_index!= shared_data[first_column_value]:
                rows_to_delete_final.append(row_index)

    for row_index in sorted(rows_to_delete_final, reverse=True):
        sheet.delete_rows(row_index)

    workbook.save(output_file_path)


if __name__ == '__main__':
    process_sheet_data_parallel()